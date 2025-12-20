#!/usr/bin/env python3
"""
Backtest Excel Report Generator - Detailed Version
Generates comprehensive Excel reports with betting odds for AI prediction backtest
"""

import json
import subprocess
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def run_backtest(start_date, end_date, limit):
    """Run backtest via the API endpoint"""
    import urllib.request
    import urllib.error

    url = f"http://localhost:3001/matches/backtest"
    data = json.dumps({
        "startDate": start_date,
        "endDate": end_date,
        "limit": limit
    }).encode('utf-8')

    headers = {
        'Content-Type': 'application/json'
    }

    try:
        req = urllib.request.Request(url, data=data, headers=headers, method='POST')
        with urllib.request.urlopen(req, timeout=300) as response:
            return json.loads(response.read().decode('utf-8'))
    except urllib.error.URLError as e:
        print(f"Error connecting to API: {e}")
        print("Make sure the server is running on port 3001")
        return None
    except Exception as e:
        print(f"Error: {e}")
        return None

def format_odds(value):
    """Format odds value"""
    if value is None or value == 0:
        return "-"
    return f"{value:.2f}"

def create_detailed_matches_sheet(wb, matches, sheet_name="Detailed Results"):
    """Create detailed matches sheet with all data per match"""
    ws = wb.create_sheet(title=sheet_name)

    current_row = 1

    for idx, match_data in enumerate(matches, 1):
        m = match_data.get('match', {})
        p = match_data.get('prediction', {})
        r = match_data.get('results', {})
        odds = p.get('odds', {}) or {}

        # Match Header
        ws.cell(row=current_row, column=1, value=f"MATCH {idx}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.cell(row=current_row, column=1).font = HEADER_FONT
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        current_row += 1

        # Match Info
        kickoff = m.get('kickoffTime', '')
        if kickoff:
            try:
                dt = datetime.fromisoformat(kickoff.replace('Z', '+00:00'))
                kickoff = dt.strftime('%Y-%m-%d %H:%M')
            except:
                pass

        info_data = [
            ("Date", kickoff),
            ("League", m.get('league', 'Unknown')),
            ("Home Team", m.get('homeTeam', '')),
            ("Away Team", m.get('awayTeam', '')),
            ("Final Score", f"{m.get('homeScore', 0)} - {m.get('awayScore', 0)}"),
            ("Total Goals", m.get('totalGoals', 0)),
        ]

        ws.cell(row=current_row, column=1, value="MATCH INFO").font = BOLD_FONT
        ws.cell(row=current_row, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        for label, value in info_data:
            ws.cell(row=current_row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

        current_row += 1

        # Betting Odds Section
        ws.cell(row=current_row, column=1, value="BETTING ODDS").font = BOLD_FONT
        ws.cell(row=current_row, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        # 1X2 Odds
        ws.cell(row=current_row, column=1, value="1X2 Odds:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=f"1: {format_odds(odds.get('homeWin'))}")
        ws.cell(row=current_row, column=3, value=f"X: {format_odds(odds.get('draw'))}")
        ws.cell(row=current_row, column=4, value=f"2: {format_odds(odds.get('awayWin'))}")
        current_row += 1

        # BTTS Odds
        ws.cell(row=current_row, column=1, value="BTTS Odds:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=f"Yes: {format_odds(odds.get('bttsYes'))}")
        ws.cell(row=current_row, column=3, value=f"No: {format_odds(odds.get('bttsNo'))}")
        current_row += 1

        # O/U 2.5 Odds
        ws.cell(row=current_row, column=1, value="O/U 2.5 Odds:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=f"Over: {format_odds(odds.get('over25'))}")
        ws.cell(row=current_row, column=3, value=f"Under: {format_odds(odds.get('under25'))}")
        current_row += 1

        current_row += 1

        # Our Predictions Section
        ws.cell(row=current_row, column=1, value="OUR PREDICTIONS").font = BOLD_FONT
        ws.cell(row=current_row, column=1).fill = SUBHEADER_FILL
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1

        # Match Outcome Prediction
        ws.cell(row=current_row, column=1, value="Match Outcome:").font = BOLD_FONT
        pred_outcome = p.get('predictedOutcome', '')
        actual_outcome = m.get('actualOutcome', '')
        outcome_correct = r.get('outcomeCorrect', False)

        outcome_cell = ws.cell(row=current_row, column=2,
                               value=f"Predicted: {pred_outcome} | Actual: {actual_outcome}")
        ws.cell(row=current_row, column=3, value="CORRECT" if outcome_correct else "WRONG")
        ws.cell(row=current_row, column=3).fill = GREEN_FILL if outcome_correct else RED_FILL
        ws.cell(row=current_row, column=3).font = BOLD_FONT
        current_row += 1

        # Probabilities
        ws.cell(row=current_row, column=1, value="Probabilities:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=f"1: {p.get('homeWinProb', 0):.1f}%")
        ws.cell(row=current_row, column=3, value=f"X: {p.get('drawProb', 0):.1f}%")
        ws.cell(row=current_row, column=4, value=f"2: {p.get('awayWinProb', 0):.1f}%")
        current_row += 1

        ws.cell(row=current_row, column=1, value="Confidence:").font = BOLD_FONT
        conf = p.get('outcomeConfidence', 0)
        conf_cell = ws.cell(row=current_row, column=2, value=f"{conf}%")
        if conf >= 70:
            conf_cell.fill = GREEN_FILL
        elif conf >= 50:
            conf_cell.fill = YELLOW_FILL
        else:
            conf_cell.fill = RED_FILL
        current_row += 1

        # BTTS Prediction
        ws.cell(row=current_row, column=1, value="BTTS:").font = BOLD_FONT
        pred_btts = p.get('predictedBTTS', '')
        actual_btts = "Yes" if m.get('actualBTTS') else "No"
        btts_correct = r.get('bttsCorrect', False)

        ws.cell(row=current_row, column=2, value=f"Predicted: {pred_btts} ({p.get('bttsYesProb', 0):.1f}% Yes)")
        ws.cell(row=current_row, column=3, value=f"Actual: {actual_btts}")
        ws.cell(row=current_row, column=4, value="CORRECT" if btts_correct else "WRONG")
        ws.cell(row=current_row, column=4).fill = GREEN_FILL if btts_correct else RED_FILL
        ws.cell(row=current_row, column=4).font = BOLD_FONT
        current_row += 1

        # Over/Under 2.5 Prediction
        ws.cell(row=current_row, column=1, value="Over/Under 2.5:").font = BOLD_FONT
        pred_over = "Over" if p.get('predictedOver25') else "Under"
        actual_over = "Over" if m.get('actualOver25') else "Under"
        over_correct = r.get('over25Correct', False)

        ws.cell(row=current_row, column=2, value=f"Predicted: {pred_over} ({p.get('over25Prob', 0):.1f}%)")
        ws.cell(row=current_row, column=3, value=f"Actual: {actual_over}")
        ws.cell(row=current_row, column=4, value="CORRECT" if over_correct else "WRONG")
        ws.cell(row=current_row, column=4).fill = GREEN_FILL if over_correct else RED_FILL
        ws.cell(row=current_row, column=4).font = BOLD_FONT
        current_row += 1

        # Expected Goals
        ws.cell(row=current_row, column=1, value="Expected Goals:").font = BOLD_FONT
        xg = p.get('expectedGoals', 0)
        actual_goals = m.get('totalGoals', 0)
        xg_error = abs(xg - actual_goals)
        ws.cell(row=current_row, column=2, value=f"xG: {xg:.2f}")
        ws.cell(row=current_row, column=3, value=f"Actual: {actual_goals}")
        ws.cell(row=current_row, column=4, value=f"Error: {xg_error:.2f}")
        current_row += 1

        # Predicted Score
        ws.cell(row=current_row, column=1, value="Predicted Score:").font = BOLD_FONT
        pred_score = p.get('predictedScore', '')
        actual_score = f"{m.get('homeScore', 0)}-{m.get('awayScore', 0)}"
        score_correct = r.get('scoreCorrect', False)
        ws.cell(row=current_row, column=2, value=f"Predicted: {pred_score}")
        ws.cell(row=current_row, column=3, value=f"Actual: {actual_score}")
        if score_correct:
            ws.cell(row=current_row, column=4, value="CORRECT!")
            ws.cell(row=current_row, column=4).fill = GREEN_FILL
            ws.cell(row=current_row, column=4).font = BOLD_FONT
        current_row += 1

        # Summary row
        current_row += 1
        total_correct = sum([outcome_correct, btts_correct, over_correct])
        ws.cell(row=current_row, column=1, value="MATCH SUMMARY:").font = BOLD_FONT
        summary_cell = ws.cell(row=current_row, column=2, value=f"{total_correct}/3 predictions correct")
        if total_correct == 3:
            summary_cell.fill = GREEN_FILL
        elif total_correct >= 2:
            summary_cell.fill = YELLOW_FILL
        else:
            summary_cell.fill = RED_FILL

        # Separator
        current_row += 2
        ws.cell(row=current_row, column=1, value="─" * 80)
        current_row += 2

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    return ws

def create_summary_table_sheet(wb, matches, sheet_name="Summary Table"):
    """Create a compact summary table with all matches"""
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    headers = [
        "No", "Date", "League", "Home", "Away", "Score",
        "Pred", "Actual", "1X2 OK",
        "Home%", "Draw%", "Away%", "Conf%",
        "1 Odds", "X Odds", "2 Odds",
        "BTTS Pred", "BTTS Act", "BTTS OK",
        "O2.5 Pred", "O2.5 Act", "O2.5 OK",
        "xG", "Goals"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, match_data in enumerate(matches, 2):
        m = match_data.get('match', {})
        p = match_data.get('prediction', {})
        r = match_data.get('results', {})
        odds = p.get('odds', {}) or {}

        # Format date
        kickoff = m.get('kickoffTime', '')
        if kickoff:
            try:
                dt = datetime.fromisoformat(kickoff.replace('Z', '+00:00'))
                kickoff = dt.strftime('%m/%d')
            except:
                kickoff = kickoff[:10]

        row_data = [
            row_idx - 1,
            kickoff,
            (m.get('league', '')[:15] + '..') if len(m.get('league', '')) > 15 else m.get('league', ''),
            m.get('homeTeam', '')[:12],
            m.get('awayTeam', '')[:12],
            f"{m.get('homeScore', 0)}-{m.get('awayScore', 0)}",
            p.get('predictedOutcome', ''),
            m.get('actualOutcome', ''),
            "Y" if r.get('outcomeCorrect') else "N",
            round(p.get('homeWinProb', 0), 1),
            round(p.get('drawProb', 0), 1),
            round(p.get('awayWinProb', 0), 1),
            round(p.get('outcomeConfidence', 0), 0),
            format_odds(odds.get('homeWin')),
            format_odds(odds.get('draw')),
            format_odds(odds.get('awayWin')),
            p.get('predictedBTTS', ''),
            "Y" if m.get('actualBTTS') else "N",
            "Y" if r.get('bttsCorrect') else "N",
            "O" if p.get('predictedOver25') else "U",
            "O" if m.get('actualOver25') else "U",
            "Y" if r.get('over25Correct') else "N",
            round(p.get('expectedGoals', 0), 1),
            m.get('totalGoals', 0)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

            # Color correct/incorrect cells
            if col == 9:  # 1X2 OK
                cell.fill = GREEN_FILL if value == "Y" else RED_FILL
            elif col == 19:  # BTTS OK
                cell.fill = GREEN_FILL if value == "Y" else RED_FILL
            elif col == 22:  # O2.5 OK
                cell.fill = GREEN_FILL if value == "Y" else RED_FILL
            elif col == 13:  # Confidence
                if value >= 70:
                    cell.fill = GREEN_FILL
                elif value >= 50:
                    cell.fill = YELLOW_FILL

    # Adjust column widths
    widths = [4, 6, 15, 12, 12, 5, 5, 5, 5, 6, 6, 6, 5, 6, 6, 6, 6, 5, 5, 6, 5, 5, 4, 5]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws

def create_statistics_sheet(wb, report, match_count):
    """Create statistics summary sheet"""
    ws = wb.create_sheet(title="Statistics", index=0)

    accuracy = report.get('accuracy', {})
    profitability = report.get('profitability', {})
    summary = report.get('summary', {})

    # Title
    ws['A1'] = f"BACKTEST REPORT - {match_count} MATCHES"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    # Report Info
    date_range = summary.get('dateRange', {})
    ws['A3'] = f"Period: {date_range.get('from', 'N/A')} to {date_range.get('to', 'N/A')}"
    ws['A4'] = f"Total Matches Analyzed: {summary.get('totalMatches', 0)}"
    ws['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Accuracy Summary Table
    ws['A7'] = "ACCURACY SUMMARY"
    ws['A7'].font = Font(bold=True, size=14)

    # Headers
    headers = ["Prediction Type", "Correct", "Total", "Accuracy %", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    outcome = accuracy.get('matchOutcome', {})
    btts = accuracy.get('btts', {})
    over_under = accuracy.get('overUnder', {})

    data = [
        ["Match Outcome (1X2)", outcome.get('correct', 0), outcome.get('total', 0), outcome.get('percentage', 0)],
        ["BTTS", btts.get('correct', 0), btts.get('total', 0), btts.get('percentage', 0)],
        ["Over/Under 2.5", over_under.get('over25', {}).get('correct', 0), over_under.get('over25', {}).get('total', 0), over_under.get('over25', {}).get('percentage', 0)],
        ["Over/Under 1.5", over_under.get('over15', {}).get('correct', 0), over_under.get('over15', {}).get('total', 0), over_under.get('over15', {}).get('percentage', 0)],
        ["Over/Under 3.5", over_under.get('over35', {}).get('correct', 0), over_under.get('over35', {}).get('total', 0), over_under.get('over35', {}).get('percentage', 0)],
    ]

    for row_idx, row_data in enumerate(data, 9):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

        # Status column
        pct = row_data[3]
        status_cell = ws.cell(row=row_idx, column=5)
        if pct >= 60:
            status_cell.value = "GOOD"
            status_cell.fill = GREEN_FILL
        elif pct >= 50:
            status_cell.value = "OK"
            status_cell.fill = YELLOW_FILL
        else:
            status_cell.value = "POOR"
            status_cell.fill = RED_FILL
        status_cell.border = THIN_BORDER

    # Confidence Breakdown
    ws['A16'] = "ACCURACY BY CONFIDENCE"
    ws['A16'].font = Font(bold=True, size=14)

    conf_headers = ["Level", "Correct", "Total", "Accuracy %"]
    for col, h in enumerate(conf_headers, 1):
        cell = ws.cell(row=17, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    by_conf = outcome.get('byConfidence', {})
    conf_data = [
        ["High (>=70%)", by_conf.get('high', {}).get('correct', 0), by_conf.get('high', {}).get('total', 0), by_conf.get('high', {}).get('percentage', 0)],
        ["Medium (50-69%)", by_conf.get('medium', {}).get('correct', 0), by_conf.get('medium', {}).get('total', 0), by_conf.get('medium', {}).get('percentage', 0)],
        ["Low (<50%)", by_conf.get('low', {}).get('correct', 0), by_conf.get('low', {}).get('total', 0), by_conf.get('low', {}).get('percentage', 0)],
    ]

    for row_idx, row_data in enumerate(conf_data, 18):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col == 4 and value > 0:
                if value >= 60:
                    cell.fill = GREEN_FILL
                elif value >= 50:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = RED_FILL

    # xG Analysis
    ws['A23'] = "EXPECTED GOALS ANALYSIS"
    ws['A23'].font = Font(bold=True, size=14)

    xg = accuracy.get('expectedGoals', {})
    ws['A24'] = f"Average xG Error: {xg.get('avgError', 'N/A')} goals"
    ws['A25'] = f"Correlation: {xg.get('correlation', 'N/A')}"

    # Insights
    insights = report.get('insights', [])
    if insights:
        ws['A28'] = "KEY INSIGHTS"
        ws['A28'].font = Font(bold=True, size=14)
        for i, insight in enumerate(insights, 29):
            ws.cell(row=i, column=1, value=f"• {insight}")

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10

    return ws

def generate_excel_report(report, match_count, output_file):
    """Generate Excel report from backtest data"""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    matches = report.get('matches', [])

    # Create sheets
    create_statistics_sheet(wb, report, match_count)
    create_summary_table_sheet(wb, matches)
    create_detailed_matches_sheet(wb, matches)

    # Save
    wb.save(output_file)
    print(f"Report saved: {output_file}")

def main():
    print("=" * 60)
    print("AI Prediction Backtest - Excel Report Generator")
    print("=" * 60)

    # Use 2024-2025 season date range
    start_str = "2024-10-01"
    end_str = "2024-12-15"

    print(f"\nDate Range: {start_str} to {end_str}")
    print("\nGenerating reports for 20, 50, and 100 matches...")

    # Generate reports for different match counts
    match_counts = [20, 50, 100]

    for count in match_counts:
        print(f"\n--- Running {count}-match backtest ---")

        report = run_backtest(start_str, end_str, count)

        if report:
            output_file = f"backtest_report_{count}_matches.xlsx"
            generate_excel_report(report, count, output_file)

            # Print quick summary
            accuracy = report.get('accuracy', {})
            outcome = accuracy.get('matchOutcome', {})
            btts = accuracy.get('btts', {})
            over25 = accuracy.get('overUnder', {}).get('over25', {})

            print(f"  Matches: {report.get('summary', {}).get('totalMatches', 0)}")
            print(f"  1X2 Accuracy: {outcome.get('percentage', 0)}%")
            print(f"  BTTS Accuracy: {btts.get('percentage', 0)}%")
            print(f"  O/U 2.5 Accuracy: {over25.get('percentage', 0)}%")
        else:
            print(f"  Failed to generate {count}-match report")

    print("\n" + "=" * 60)
    print("Report generation complete!")
    print("=" * 60)

if __name__ == "__main__":
    main()
